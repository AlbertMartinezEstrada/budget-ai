import openpyxl
from datetime import datetime

def guardar_a_excel(df_classificat, excel_path):
    # 1. Carreguem el llibre d'estil
    wb = openpyxl.load_workbook(excel_path)
    
    # Diccionari per traduir número de mes a nom de pestanya (en català com tu ho tens)
    mesos = {
        1: "GENER", 2: "FEBRER", 3: "MARÇ", 4: "ABRIL", 
        5: "MAIG", 6: "JUNY", 7: "JULIOL", 8: "AGOST", 
        9: "SETEMBRE", 10: "OCTUBRE", 11: "NOVEMBRE", 12: "DESEMBRE"
    }

    for _, fila in df_classificat.iterrows():
        # Obtenir el mes de la data (format dd/mm/yyyy)
        data_dt = datetime.strptime(fila['date'], '%d/%m/%Y')
        nom_pestanya = mesos[data_dt.month]
        
        if nom_pestanya in wb.sheetnames:
            ws = wb[nom_pestanya]
            
            # Busquem la secció de GASTOS VARIABLES
            # Segons el teu fitxer, la capçalera està al voltant de la fila 10, columna I
            # Anem a buscar la primera fila lliure a partir de la columna I (9)
            start_row = 11
            while ws.cell(row=start_row, column=9).value is not None:
                start_row += 1
            
            # Escrivim les dades en l'ordre de les teves columnes:
            # I: Empresa, J: Descripcio, K: Tipus, L: Mètode, M: Cost
            ws.cell(row=start_row, column=9).value = fila['companyName']
            ws.cell(row=start_row, column=10).value = fila['description_curta']
            ws.cell(row=start_row, column=11).value = fila['category']
            ws.cell(row=start_row, column=12).value = "La Caixa" # Mètode per defecte
            ws.cell(row=start_row, column=13).value = float(str(fila['cost']).replace(',', '.'))
            
    # Guardem els canvis
    wb.save(excel_path)
    print(f"✅ Excel actualitzat correctament!")